"""将 APP 功能遍历结果导出为 Excel。"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

REGION_LABELS = {
    "top_tab": "顶部 Tab",
    "top": "顶部",
    "bottom_tab": "底部 Tab",
    "bottom": "底部",
    "side": "侧栏",
    "left": "左侧",
    "right": "右侧",
    "button": "按钮",
    "list_item": "列表项",
    "other": "其他",
}

STATUS_LABELS = {
    "listed": "已列出",
    "visited": "已访问",
}

NODE_TYPE_LABELS = {
    "application": "应用",
    "screen": "功能",
    "module": "功能",
    "function": "功能",
}


def _level_columns(path: list[str], max_levels: int = 5) -> dict[str, str]:
    out: dict[str, str] = {}
    for i in range(max_levels):
        key = f"level_{i + 1}"
        out[key] = path[i] if i < len(path) else ""
    return out


def build_feature_workbook(
    tree: dict[str, Any],
    *,
    device_id: str = "",
    model_name: str = "",
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "功能清单"

    headers = [
        "序号",
        "功能类型",
        "功能点名称",
        "功能点描述",
        "位置信息",
        "一级功能",
        "二级功能",
        "三级功能",
        "四级功能",
        "五级功能",
        "完整路径",
        "层级",
        "区域",
        "页面标题",
        "发现状态",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    features = tree.get("features") or []
    if not isinstance(features, list):
        features = []

    for idx, raw in enumerate(features, start=1):
        if not isinstance(raw, dict):
            continue
        path = raw.get("path") or []
        if not isinstance(path, list):
            path = [str(path)]
        path = [str(p) for p in path]
        levels = _level_columns(path)
        full_path = " > ".join(path) if path else ""
        region = str(raw.get("region") or "")
        status = str(raw.get("status") or "listed")
        name = str(raw.get("name") or (path[-1] if path else ""))
        ws.append(
            [
                idx,
                raw.get("function_type") or REGION_LABELS.get(region, region or "—"),
                name,
                raw.get("description") or "",
                raw.get("location") or full_path,
                levels["level_1"],
                levels["level_2"],
                levels["level_3"],
                levels["level_4"],
                levels["level_5"],
                full_path,
                raw.get("depth", len(path)),
                REGION_LABELS.get(region, region or "—"),
                raw.get("screen_title") or "",
                STATUS_LABELS.get(status, status),
            ]
        )

    meta = wb.create_sheet("元数据")
    meta.append(["项目", "值"])
    meta["A1"].font = Font(bold=True)
    meta["B1"].font = Font(bold=True)
    rows = [
        ("APP 名称", tree.get("app_name") or ""),
        ("Bundle ID", tree.get("bundle_id") or ""),
        ("设备 ID", device_id),
        ("模型", model_name),
        ("开始时间", tree.get("started_at") or ""),
        ("结束时间", tree.get("finished_at") or ""),
        ("访问页面数", tree.get("screens_visited") or 0),
        ("功能项总数", len(features)),
        ("导出时间", datetime.utcnow().isoformat(timespec="seconds") + "Z"),
    ]
    for r in rows:
        meta.append(list(r))

    return wb


def write_explore_excel(
    tree: dict[str, Any],
    dest: Path,
    *,
    device_id: str = "",
    model_name: str = "",
) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = build_feature_workbook(tree, device_id=device_id, model_name=model_name)
    wb.save(dest)
    return dest


def workbook_to_bytes(
    tree: dict[str, Any],
    *,
    device_id: str = "",
    model_name: str = "",
) -> bytes:
    buf = io.BytesIO()
    wb = build_feature_workbook(tree, device_id=device_id, model_name=model_name)
    wb.save(buf)
    return buf.getvalue()


def _full_path_for_tree_node(node: dict[str, Any]) -> str:
    if not isinstance(node, dict):
        return ""
    if node.get("node_type") == "application":
        return str(node.get("name") or "").strip()
    prefix = [str(p).strip() for p in (node.get("path") or []) if str(p).strip()]
    name = str(node.get("name") or "").strip()
    if name and (not prefix or prefix[-1] != name):
        prefix = [*prefix, name]
    return " > ".join(prefix)


def _function_type_for_tree_node(node: dict[str, Any]) -> str:
    if node.get("node_type") == "application":
        return "应用"
    ft = str(node.get("function_type") or "").strip()
    if ft:
        return ft
    region = str(node.get("region") or "")
    return REGION_LABELS.get(region, region or "功能")


def _iter_tree_rows(node: dict[str, Any] | None, depth: int = 0) -> list[list[Any]]:
    if not isinstance(node, dict):
        return []
    rows: list[list[Any]] = []
    full_path = _full_path_for_tree_node(node)
    node_type = str(node.get("node_type") or "")
    rows.append(
        [
            depth,
            NODE_TYPE_LABELS.get(node_type, node_type or "—"),
            str(node.get("name") or "").strip(),
            _function_type_for_tree_node(node),
            full_path,
            str(node.get("description") or "").strip(),
            str(node.get("location") or full_path).strip(),
        ]
    )
    for ch in node.get("children") or []:
        if isinstance(ch, dict):
            rows.extend(_iter_tree_rows(ch, depth + 1))
    return rows


def _resolve_function_tree_root(tree: dict[str, Any]) -> dict[str, Any] | None:
    ft = tree.get("function_tree") or tree.get("function_tree_by_path")
    if isinstance(ft, dict):
        return ft
    app_name = str(tree.get("app_name") or "应用").strip() or "应用"
    features = tree.get("features") or []
    if not isinstance(features, list):
        features = []
    from app.services.agent_service_client import build_function_tree

    return build_function_tree(app_name, features)


def build_feature_tree_detail_workbook(tree: dict[str, Any]) -> Workbook:
    """已确认功能树导出：功能点信息 + 功能清单树 两个 sheet。"""
    wb = Workbook()
    ws_points = wb.active
    ws_points.title = "功能点信息"

    point_headers = ["序号", "功能类型", "功能点名称", "功能点描述", "位置信息"]
    ws_points.append(point_headers)
    for cell in ws_points[1]:
        cell.font = Font(bold=True)

    features = tree.get("features") or []
    if not isinstance(features, list):
        features = []

    for idx, raw in enumerate(features, start=1):
        if not isinstance(raw, dict):
            continue
        path = raw.get("path") or []
        if not isinstance(path, list):
            path = [str(path)]
        path = [str(p) for p in path if str(p).strip()]
        region = str(raw.get("region") or "")
        name = str(raw.get("name") or (path[-1] if path else ""))
        full_path = " > ".join(path) if path else ""
        ws_points.append(
            [
                idx,
                raw.get("function_type") or REGION_LABELS.get(region, region or "—"),
                name,
                raw.get("description") or "",
                raw.get("location") or full_path,
            ]
        )

    ws_tree = wb.create_sheet("功能清单树")
    tree_headers = [
        "序号",
        "层级",
        "节点类型",
        "功能点名称",
        "功能类型",
        "完整路径",
        "功能点描述",
        "位置信息",
    ]
    ws_tree.append(tree_headers)
    for cell in ws_tree[1]:
        cell.font = Font(bold=True)

    root = _resolve_function_tree_root(tree)
    flat = _iter_tree_rows(root, 0) if root else []
    for seq, row in enumerate(flat, start=1):
        ws_tree.append([seq, *row])

    return wb


def tree_detail_workbook_to_bytes(tree: dict[str, Any]) -> bytes:
    buf = io.BytesIO()
    wb = build_feature_tree_detail_workbook(tree)
    wb.save(buf)
    return buf.getvalue()
