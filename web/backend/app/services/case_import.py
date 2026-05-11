"""从 Excel / CSV 批量导入测试用例（宽松列名匹配）。"""

from __future__ import annotations

import csv
import io
import json
import re
from typing import Any

from pydantic import ValidationError

from app.schemas import CaseStepJson, TestCaseCreate


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace("\ufeff", "")


def _pick(row: dict[str, Any], *aliases: str) -> str:
    keys = {_norm_header(k): v for k, v in row.items()}
    for a in aliases:
        nk = _norm_header(a)
        if nk in keys and keys[nk] is not None:
            return str(keys[nk]).strip()
    return ""


_STEP_LINE_RE = re.compile(r"^\s*(\d+)\s*[\.\)、]\s*(.+)$")


def _steps_from_cell(cell: str) -> list[CaseStepJson]:
    """支持 JSON 数组或按行「1. xxx / 1、xxx」。"""
    cell = (cell or "").strip()
    if not cell:
        return []
    if cell.startswith("["):
        try:
            data = json.loads(cell)
            if isinstance(data, list):
                out: list[CaseStepJson] = []
                for i, item in enumerate(data, start=1):
                    if isinstance(item, dict):
                        out.append(
                            CaseStepJson(
                                order=int(item.get("order", i)),
                                description=str(item.get("description", "")),
                                expected=str(item.get("expected", "")),
                            )
                        )
                    elif isinstance(item, str):
                        out.append(CaseStepJson(order=i, description=item, expected=""))
                return out
        except json.JSONDecodeError:
            pass
    lines = [ln.strip() for ln in cell.splitlines() if ln.strip()]
    steps: list[CaseStepJson] = []
    for i, ln in enumerate(lines, start=1):
        m = _STEP_LINE_RE.match(ln)
        if m:
            steps.append(CaseStepJson(order=int(m.group(1)), description=m.group(2).strip(), expected=""))
        else:
            steps.append(CaseStepJson(order=i, description=ln, expected=""))
    return steps


def rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    import io as bio

    from openpyxl import load_workbook

    wb = load_workbook(bio.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header]
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        item: dict[str, Any] = {}
        for i, key in enumerate(headers):
            if not key:
                continue
            item[key] = row[i] if i < len(row) else None
        if any(v not in (None, "") for v in item.values()):
            out.append(item)
    return out


def parse_import_file(content: bytes, filename: str) -> tuple[list[dict[str, Any]], list[str]]:
    fn = (filename or "").lower()
    errs: list[str] = []
    try:
        if fn.endswith(".csv"):
            rows = rows_from_csv(content)
        elif fn.endswith(".xlsx") or fn.endswith(".xls"):
            rows = rows_from_xlsx(content)
        else:
            return [], ["仅支持 .csv 或 .xlsx 文件"]
    except Exception as e:
        return [], [f"解析失败：{e}"]
    return rows, errs


def row_to_create(project_id: int, row: dict[str, Any]) -> TestCaseCreate | None:
    title = _pick(row, "title", "标题", "用例标题", "名称", "case_title")
    task_text = _pick(row, "task_text", "任务描述", "执行说明", "说明", "描述")
    pre = _pick(row, "preconditions", "前置条件", "前置", "precondition")
    pri = _pick(row, "priority", "优先级", "priority") or "P2"
    steps_cell = _pick(row, "steps", "步骤", "测试步骤", "step")
    exp_cell = _pick(row, "expected", "预期", "预期结果", "期望")

    steps = _steps_from_cell(steps_cell)
    if exp_cell and not steps:
        steps = [CaseStepJson(order=1, description="执行用例", expected=exp_cell)]
    elif exp_cell and len(steps) == 1 and not steps[0].expected:
        steps = [CaseStepJson(order=1, description=steps[0].description, expected=exp_cell)]

    try:
        return TestCaseCreate(
            project_id=project_id,
            title=(title or "未命名用例").strip() or "未命名用例",
            task_text=task_text,
            preconditions=pre,
            steps=steps,
            priority=(pri[:16] if pri else "P2"),
        )
    except ValidationError:
        return None
